"""Batch Analysis Service (re-stabilized version).

This file was previously corrupted multiple times by stray ``self.*`` lines at
class scope. It has been rewritten from scratch to provide a compact, robust
implementation with the following features:

Core lifecycle:
    - create_job / start_job / cancel_job / update_task_progress
    - list_jobs / get_job_status / stats & pruning

Caching awareness:
    - Detects per-task cached results (via batch_result_cache service) and
        performs immediate completion for fully cached jobs (``cache_only``).

WebSocket events (best-effort):
    - batch_created, batch_started, task_progress, batch_completed,
        batch_cancelled, queue_depth_update

Priority queue manager:
    - In‑memory priority queues (high, normal, low) with FIFO ordering.
    - ``create_job`` enqueues a job (optionally with ``options['priority']``) and
        starts it immediately unless ``options['defer_start']`` is True (used by
        tests to observe queue depth deterministically).
    - ``dispatch_next()`` helper allows manual dispatch of the next queued job
        honoring priority order (high→normal→low).

Analytics:
    - ``get_job_stats`` returns counts plus avg wait / run times, cache hit rate,
        and failure rate.

Safeguards:
    - Uses ``@dataclass`` for ``BatchJob`` so attributes cannot silently drift out
        of initializer scope again.
    - All WebSocket emission is wrapped in best-effort logic; failures are logged
        at debug level only.

Future work placeholders (not yet implemented here): retries, persistence for
queue, dependency graphs, resource throttling, advanced scheduling.
"""

from __future__ import annotations

import logging
import uuid
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional
from dataclasses import dataclass, field
from collections import deque

from ..models import BatchAnalysis  # type: ignore
from ..constants import Paths  # for REPORTS_DIR persistence of export artifacts
from ..constants import JobStatus, AnalysisType  # type: ignore
from flask import has_app_context  # type: ignore
from ..extensions import get_session, db  # type: ignore
from .batch_result_cache_service import batch_result_cache, CacheKeyParts  # type: ignore
try:  # optional websocket integration marker
    from . import websocket_integration  # type: ignore  # noqa: F401
    _WS_AVAILABLE = True
except Exception:  # pragma: no cover
    _WS_AVAILABLE = False

logger = logging.getLogger(__name__)
logger.debug("batch_service loaded (stabilized version with priority queue)")


class PriorityQueueManager:
    """Simple in-memory priority queue (high > normal > low)."""

    PRIORITIES = ("high", "normal", "low")

    def __init__(self, owner: 'BatchAnalysisService') -> None:
        self._owner = owner
        self._queues: Dict[str, deque[str]] = {p: deque() for p in self.PRIORITIES}
        self._meta: Dict[str, dict] = {}

    def enqueue(self, batch_id: str, priority: str = 'normal', metadata: Optional[dict] = None) -> bool:
        p = priority if priority in self.PRIORITIES else 'normal'
        # If already queued, treat as idempotent success (API may attempt double enqueue)
        if batch_id in self._meta:
            return True
        self._queues[p].append(batch_id)
        self._meta[batch_id] = metadata or {}
        self._owner._emit('queue_depth_update', self.status_overview())
        return True

    def dequeue(self) -> Optional[str]:
        for p in self.PRIORITIES:
            q = self._queues[p]
            if q:
                bid = q.popleft()
                self._meta.pop(bid, None)
                self._owner._emit('queue_depth_update', self.status_overview())
                return bid
        return None

    def cancel(self, batch_id: str) -> bool:
        for p in self.PRIORITIES:
            q = self._queues[p]
            try:
                q.remove(batch_id)
                self._meta.pop(batch_id, None)
                self._owner._emit('queue_depth_update', self.status_overview())
                return True
            except ValueError:
                continue
        return False

    def status_overview(self) -> dict:
        depths = {p: len(self._queues[p]) for p in self.PRIORITIES}
        return {"depths": depths, "total": sum(depths.values())}

    def is_enqueued(self, batch_id: str) -> bool:
        return batch_id in self._meta


@dataclass
class BatchJob:
    """In-memory representation of a batch job.

    Converted to dataclass to avoid prior corruption where attribute assignment
    lines became detached from __init__ resulting in NameError during module
    import. Dataclass fields make initialization declarative and resilient.
    """
    id: str
    name: str
    description: str
    analysis_types: List[AnalysisType]
    models: List[str]
    app_range: List[int]
    options: Dict[str, Any] = field(default_factory=dict)
    status: JobStatus = JobStatus.PENDING
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    started_at: Optional[datetime] = None
    completed_at: Optional[datetime] = None
    total_tasks: int = 0
    completed_tasks: int = 0
    failed_tasks: int = 0
    results: Dict[str, Any] = field(default_factory=dict)
    errors: List[str] = field(default_factory=list)
    # Caching metadata
    cache_only: bool = False
    cached_tasks: int = 0
    cache_hit: bool = False

    def progress_pct(self) -> float:
        return (self.completed_tasks / self.total_tasks * 100) if self.total_tasks else 0.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            'id': self.id,
            'name': self.name,
            'description': self.description,
            'status': self.status.value,
            'analysis_types': [a.value for a in self.analysis_types],
            'models': self.models,
            'app_range': self.app_range,
            'total_tasks': self.total_tasks,
            'completed_tasks': self.completed_tasks,
            'failed_tasks': self.failed_tasks,
            'progress_percentage': self.progress_pct(),
            'cache_only': self.cache_only,
            'cached_tasks': self.cached_tasks,
            'cache_hit': self.cache_hit,
        }


class BatchAnalysisService:
    """Service coordinating batch job lifecycle with priority queue, caching, and events."""

    def __init__(self) -> None:
        self.logger = logger
        self.jobs: Dict[str, BatchJob] = {}
        self.queue_manager = PriorityQueueManager(self)
        try:  # pragma: no cover
            from .batch_scheduler import batch_scheduler_service  # type: ignore
            self.scheduler = batch_scheduler_service
        except Exception:  # pragma: no cover
            self.scheduler = None
        if _WS_AVAILABLE:
            try:  # pragma: no cover
                from .websocket_integration import WebSocketIntegration  # type: ignore
                self._ws = WebSocketIntegration()
            except Exception:
                self._ws = None
        else:
            self._ws = None

    # ---------------- internal helpers ---------------- #
    def _emit(self, event: str, payload: Dict[str, Any]):  # pragma: no cover
        if not self._ws:
            return
        try:
            from ..extensions import socketio  # type: ignore
            socketio.emit(event, payload)
        except Exception:
            self.logger.debug(f"WS emit failed {event}: {payload}")

    def _parse_app_range(self, app_range_str: str) -> List[int]:
        vals: List[int] = []
        for part in app_range_str.split(','):
            part = part.strip()
            if not part:
                continue
            if '-' in part:
                a, b = part.split('-', 1)
                try:
                    start = int(a)
                    end = int(b)
                    vals.extend(range(start, end + 1))
                except ValueError:
                    continue
            else:
                try:
                    vals.append(int(part))
                except ValueError:
                    continue
        return sorted(set(vals))

    # ---------------- lifecycle ---------------- #
    def create_job(self, name: str, description: str, analysis_types: List[str],
                   models: List[str], app_range_str: str,
                   options: Optional[Dict[str, Any]] = None,
                   *, enqueue_immediately: bool = True) -> str:
        options = options or {}
        # Generate a collision‑resistant job id (prior implementation using only timestamp occasionally collided in fast tests)
        job_id = f"job_{int(datetime.now(timezone.utc).timestamp()*1000)}_{uuid.uuid4().hex[:6]}"
        ats: List[AnalysisType] = []
        for at in analysis_types:
            try:
                alias = {
                    'security': 'security_backend',
                    'performance': 'performance',
                    'static': 'code_quality',
                    'code_quality': 'code_quality',
                    'ai': 'openrouter',
                    'zap': 'zap_security',  # newly added alias so user can request "zap"
                }.get(at.lower(), at)
                ats.append(AnalysisType(alias))
            except ValueError:
                self.logger.warning(f"Unknown analysis type ignored: {at}")
        app_range = self._parse_app_range(app_range_str)
        job = BatchJob(job_id, name, description, ats, models, app_range, options)
        job.total_tasks = len(ats) * len(models) * len(app_range)

        # cache accounting
        for at_enum in ats:
            for model in models:
                for app_number in app_range:
                    parts: CacheKeyParts = (at_enum.value, model, app_number, options)
                    if batch_result_cache.get_cached(parts) is not None:
                        job.cached_tasks += 1
        if job.cached_tasks:
            job.cache_hit = True

        # fully cached immediate completion
        if job.total_tasks and job.cached_tasks == job.total_tasks:
            job.cache_only = True
            job.status = JobStatus.COMPLETED
            job.started_at = job.completed_at = datetime.now(timezone.utc)
            self.jobs[job_id] = job
            self._emit('batch_created', {
                'batch_id': job_id,
                'cache_only': True,
                'cached_tasks': job.cached_tasks,
                'total_tasks': job.total_tasks
            })
            self._emit('batch_completed', {
                'batch_id': job_id,
                'status': job.status.value,
                'cache_only': True,
                'completed_tasks': job.total_tasks,
                'failed_tasks': 0
            })
            self._persist_db_row(job, ats, models, app_range, cached_only=True)
            return job_id

        self.jobs[job_id] = job
        self._emit('batch_created', {
            'batch_id': job_id,
            'status': job.status.value,
            'analysis_types': [a.value for a in ats],
            'models': models,
            'app_range': app_range,
            'total_tasks': job.total_tasks,
            'cache_hit': job.cache_hit,
            'cached_tasks': job.cached_tasks,
        })
        self._persist_db_row(job, ats, models, app_range, cached_only=False)

        priority = options.get('priority', 'normal')
        defer = bool(options.get('defer_start'))
        if enqueue_immediately and self.queue_manager.enqueue(job_id, priority=priority):
            if not defer:
                self.start_job(job_id)
        return job_id

    def start_job(self, job_id: str) -> bool:
        job = self.jobs.get(job_id)
        if not job or job.status != JobStatus.PENDING:
            return False
        job.status = JobStatus.RUNNING
        job.started_at = datetime.now(timezone.utc)
        self._emit('batch_started', {
            'batch_id': job_id,
            'started_at': job.started_at.isoformat() if job.started_at else None,
            'total_tasks': job.total_tasks,
            'cached_tasks': job.cached_tasks,
        })
        self._update_db_status(job_id, JobStatus.RUNNING, started_at=job.started_at)
        return True

    def cancel_job(self, job_id: str) -> bool:
        job = self.jobs.get(job_id)
        if not job or job.status not in (JobStatus.PENDING, JobStatus.RUNNING):
            return False
        job.status = JobStatus.CANCELLED
        job.completed_at = datetime.now(timezone.utc)
        self.queue_manager.cancel(job_id)
        self._emit('batch_cancelled', {
            'batch_id': job_id,
            'status': job.status.value,
            'completed_tasks': job.completed_tasks,
            'failed_tasks': job.failed_tasks,
        })
        self._update_db_status(job_id, JobStatus.CANCELLED, completed_at=job.completed_at)
        return True

    def update_task_progress(self, job_id: str, task_completed: bool = False, task_failed: bool = False, result: Any = None):
        job = self.jobs.get(job_id)
        if not job:
            return
        if task_completed:
            job.completed_tasks += 1
        if task_failed:
            job.failed_tasks += 1
        self._emit('task_progress', {
            'batch_id': job_id,
            'completed_tasks': job.completed_tasks,
            'failed_tasks': job.failed_tasks,
            'total_tasks': job.total_tasks,
            'progress_pct': job.progress_pct(),
        })
        if job.completed_tasks + job.failed_tasks >= job.total_tasks and job.status == JobStatus.RUNNING:
            job.status = JobStatus.COMPLETED if job.failed_tasks == 0 else JobStatus.FAILED
            job.completed_at = datetime.now(timezone.utc)
            self._emit('batch_completed', {
                'batch_id': job_id,
                'status': job.status.value,
                'completed_tasks': job.completed_tasks,
                'failed_tasks': job.failed_tasks,
                'total_tasks': job.total_tasks,
                'duration_seconds': (job.completed_at - job.started_at).total_seconds() if job.started_at and job.completed_at else None,
            })
            self._update_db_status(job_id, job.status, completed_at=job.completed_at, completed_tasks=job.completed_tasks, failed_tasks=job.failed_tasks)

    # --------------- queries / stats --------------- #
    def get_job(self, job_id: str) -> Optional[BatchJob]:
        return self.jobs.get(job_id)

    def get_job_status(self, job_id: str) -> Optional[Dict[str, Any]]:
        job = self.jobs.get(job_id)
        return job.to_dict() if job else None

    def list_jobs(self) -> List[Dict[str, Any]]:
        return [j.to_dict() for j in self.jobs.values()]

    def get_jobs_summary(self) -> Dict[str, Any]:
        total = len(self.jobs)
        by_status: Dict[str, int] = {}
        for j in self.jobs.values():
            by_status[j.status.value] = by_status.get(j.status.value, 0) + 1
        return {'total': total, 'by_status': by_status}

    def get_job_stats(self) -> Dict[str, Any]:
        summary = self.get_jobs_summary()
        counts = {k.lower(): v for k, v in summary['by_status'].items()}
        for key in ('pending', 'running', 'completed', 'failed', 'cancelled'):
            counts.setdefault(key, 0)
        total = summary['total'] or 1
        wait_times: List[float] = []
        run_times: List[float] = []
        cache_hits = 0
        failures = 0
        for job in self.jobs.values():
            if job.cache_hit:
                cache_hits += 1
            if job.status == JobStatus.FAILED:
                failures += 1
            if job.started_at:
                wait_times.append((job.started_at - job.created_at).total_seconds())
            if job.started_at and job.completed_at:
                run_times.append((job.completed_at - job.started_at).total_seconds())
        def _avg(vals: List[float]) -> float: return sum(vals)/len(vals) if vals else 0.0
        return {
            'total': summary['total'],
            'pending': counts['pending'],
            'running': counts['running'],
            'completed': counts['completed'],
            'failed': counts['failed'],
            'cancelled': counts['cancelled'],
            'avg_wait_seconds': round(_avg(wait_times), 3),
            'avg_run_seconds': round(_avg(run_times), 3),
            'cache_hit_rate': round(cache_hits / total, 3),
            'failure_rate': round(failures / total, 3),
        }

    def prune_old_jobs(self, max_age_hours: int = 24):
        cutoff = datetime.now(timezone.utc).timestamp() - max_age_hours * 3600
        to_delete = [jid for jid, j in self.jobs.items() if j.created_at.timestamp() < cutoff]
        for jid in to_delete:
            self.jobs.pop(jid, None)

    # -------- dashboard snapshot (UI aggregation) -------- #
    def get_dashboard_snapshot(self, recent_limit: int = 10) -> Dict[str, Any]:
        """Return aggregated data for the batch dashboard.

        Combines persisted BatchAnalysis rows (for historical context) with the
        in-memory job map (for up-to-the-second progress) and queue depths.
        """
        snapshot: Dict[str, Any] = {}
        try:
            from ..models import BatchAnalysis  # type: ignore
            from ..extensions import db as _db  # type: ignore
            rows = _db.session.query(BatchAnalysis).order_by(BatchAnalysis.created_at.desc()).limit(200).all()
        except Exception:  # pragma: no cover
            rows = []

        # Map DB rows into simplified structures; enrich with live progress if available
        active_statuses = {JobStatus.PENDING, JobStatus.RUNNING}
        active_batches = []
        recent_batches = []
        for r in rows:
            jid = r.batch_id
            live = self.jobs.get(jid)
            total_tasks = (live.total_tasks if live else r.total_tasks) or 0
            completed_tasks = (live.completed_tasks if live else r.completed_tasks) or 0
            failed_tasks = (live.failed_tasks if live else r.failed_tasks) or 0
            progress = round(((completed_tasks + failed_tasks)/total_tasks * 100), 1) if total_tasks else 0.0
            status = (live.status.value if live else (r.status.value if hasattr(r.status, 'value') else r.status))
            model_filter = r.get_model_filter() if hasattr(r, 'get_model_filter') else []
            app_filter = r.get_app_filter() if hasattr(r, 'get_app_filter') else []
            analysis_types = r.get_analysis_types() if hasattr(r, 'get_analysis_types') else []
            batch_entry = {
                'id': jid,
                'name': getattr(r, 'name', jid),
                'status': status,
                'analysis_type': ','.join(analysis_types) if analysis_types else 'mixed',
                'application_count': len(app_filter) * len(model_filter) if app_filter and model_filter else 0,
                'created_at': getattr(r, 'created_at', None),
                'completed_at': getattr(r, 'completed_at', None),
                'progress': progress,
                'completed_analyses': completed_tasks,
                'total_analyses': total_tasks,
                'failed_analyses': failed_tasks,
                'duration_minutes': round(((getattr(r, 'completed_at', None) or datetime.now(timezone.utc)) - getattr(r, 'created_at', datetime.now(timezone.utc))).total_seconds()/60,1) if getattr(r,'created_at',None) else None,
                'completion_time': getattr(r, 'completed_at', None),
            }
            if live and live.status in active_statuses:
                active_batches.append(batch_entry)
            recent_batches.append(batch_entry)

        # Queue summary
        queue_overview = {'total': 0, 'high': {'depth': 0}, 'normal': {'depth': 0}, 'low': {'depth': 0}}
        if hasattr(self, 'queue_manager') and self.queue_manager:
            try:
                qm = self.queue_manager.status_overview() or {}
                # Normalize keys expected by template
                queue_overview['total'] = qm.get('total', 0)
                for k in ('high','normal','low'):
                    v = qm.get(k)
                    if isinstance(v, dict):
                        queue_overview[k]['depth'] = v.get('depth', 0)
                    elif hasattr(v, 'depth'):
                        queue_overview[k]['depth'] = getattr(v, 'depth', 0)
            except Exception:
                pass

        # Basic stats summary (will enrich with live worker metrics below)
        stats = {
            'total_batches': len(recent_batches),
            'running_batches': len([b for b in active_batches if b['status'] == JobStatus.RUNNING.value]),
            'queued_batches': len([b for b in active_batches if b['status'] == JobStatus.PENDING.value]),
            'completed_batches': len([b for b in recent_batches if b['status'] == JobStatus.COMPLETED.value]),
            'total_analyses': sum(b['total_analyses'] for b in recent_batches),
            'active_workers': 0,
            'registered_workers': 0,
            'busy_workers': 0,
            'scheduled_tasks': 0,
        }

        # Attempt to gather real Celery worker metrics (best-effort; never raise)
        # We avoid importing Celery globally if not installed or app not initialized.
        try:  # pragma: no cover - environment dependent
            from app.extensions import get_celery  # type: ignore
            celery_app = get_celery()
            insp = None
            if celery_app:
                try:
                    insp = celery_app.control.inspect(timeout=0.5)  # short timeout to prevent blocking request thread
                except Exception:
                    insp = None
            if insp:
                # active() returns dict: worker -> list[tasks]
                active_map = insp.active() or {}
                stats['busy_workers'] = sum(1 for _w, tasks in active_map.items() if tasks)
                stats['active_workers'] = len(active_map)  # workers responding to active probe
                # registered() returns dict: worker -> list[task_names]
                reg_map = insp.registered() or {}
                if reg_map:
                    stats['registered_workers'] = len(reg_map)
                # scheduled() returns dict: worker -> list[entries]
                sched_map = insp.scheduled() or {}
                if sched_map:
                    stats['scheduled_tasks'] = sum(len(v) for v in sched_map.values())
        except Exception:
            pass  # Silent fallback; keep zeros

        # Derive max_workers heuristic: prefer registered_workers if >0 else active_workers
        max_workers_est = max(stats.get('registered_workers', 0), stats.get('active_workers', 0), 1)

        system_resources = {
            'cpu_usage': 0.0,          # TODO: integrate psutil in future if desired
            'memory_usage': 0.0,       # TODO: integrate psutil in future
            'active_workers': stats.get('active_workers', 0),
            'max_workers': max_workers_est,
            'busy_workers': stats.get('busy_workers', 0),
            'scheduled_tasks': stats.get('scheduled_tasks', 0),
        }

        snapshot['stats'] = stats
        snapshot['active_batches'] = active_batches
        snapshot['recent_batches'] = recent_batches[:recent_limit]
        snapshot['queue_overview'] = queue_overview
        snapshot['system_resources'] = system_resources
        return snapshot

    def cleanup_old_jobs(self, max_age_hours: int = 24) -> int:
        before = len(self.jobs)
        self.prune_old_jobs(max_age_hours)
        return before - len(self.jobs)

    def dispatch_next(self) -> Optional[str]:
        bid = self.queue_manager.dequeue()
        if bid:
            self.start_job(bid)
        return bid

    # -------- test-only utilities (not part of public runtime API) -------- #
    def _reset_for_test(self):  # pragma: no cover - invoked explicitly in tests
        """Clear in-memory state so tests start from a clean slate.

        This avoids cross-test interference where prior queued items inflate
        queue depth counts for subsequent tests.
        """
        self.jobs.clear()
        # Rebuild queue manager to clear internal deques & metadata
        self.queue_manager = PriorityQueueManager(self)

    # persistence helpers
    def _persist_db_row(self, job: BatchJob, ats: List[AnalysisType], models: List[str], app_range: List[int], cached_only: bool):  # pragma: no cover
        try:
            payload = {
                'batch_id': job.id,
                'status': job.status,
                'total_tasks': job.total_tasks,
                'completed_tasks': job.completed_tasks,
                'failed_tasks': job.failed_tasks,
                'started_at': job.started_at,
                'completed_at': job.completed_at,
                'analysis_types': [a.value for a in ats],
                'model_filter': models,
                'app_filter': app_range,
                'config': {'cache_only': cached_only, **job.options},
            }
            if has_app_context():
                ba = BatchAnalysis()
                ba.batch_id = payload['batch_id']
                ba.status = payload['status']
                ba.total_tasks = payload['total_tasks']
                ba.completed_tasks = payload['completed_tasks']
                ba.failed_tasks = payload['failed_tasks']
                ba.started_at = payload['started_at']
                ba.completed_at = payload['completed_at']
                ba.set_analysis_types(payload['analysis_types'])
                ba.set_model_filter(payload['model_filter'])
                ba.set_app_filter(payload['app_filter'])
                ba.set_config(payload['config'])
                db.session.add(ba)
                db.session.commit()
            else:
                with get_session() as session:
                    ba = BatchAnalysis()
                    ba.batch_id = payload['batch_id']
                    ba.status = payload['status']
                    ba.total_tasks = payload['total_tasks']
                    ba.completed_tasks = payload['completed_tasks']
                    ba.failed_tasks = payload['failed_tasks']
                    ba.started_at = payload['started_at']
                    ba.completed_at = payload['completed_at']
                    ba.set_analysis_types(payload['analysis_types'])
                    ba.set_model_filter(payload['model_filter'])
                    ba.set_app_filter(payload['app_filter'])
                    ba.set_config(payload['config'])
                    session.add(ba)
                    session.commit()
        except Exception:
            pass

    def _update_db_status(self, job_id: str, status: JobStatus, **fields):  # pragma: no cover
        try:
            with get_session() as session:
                row = session.query(BatchAnalysis).filter_by(batch_id=job_id).first()
                if row:
                    row.status = status
                    for k, v in fields.items():
                        setattr(row, k, v)
                    session.commit()
        except Exception:
            pass

    def resume_batch(self, job_id: str) -> bool:
        return False

    # ---------------- export helpers ---------------- #
    def generate_job_export(self, job_id: str, fmt: str = 'json') -> tuple[str, str, bytes]:
        """Generate export for a single job.

        Returns (mimetype, filename, data_bytes).
        Supported fmt: json, csv, txt (alias summary), ndjson (task stream placeholder),
        xlsx/pdf are placeholders returning text for now.
        """
        fmt = (fmt or 'json').lower()
        job = self.jobs.get(job_id)
        # Attempt to hydrate from DB if not in memory
        if not job:
            try:  # pragma: no cover
                with get_session() as session:
                    row = session.query(BatchAnalysis).filter_by(batch_id=job_id).first()
                    if row:
                        # minimal reconstruction
                        ats = []
                        try:
                            for a in row.get_analysis_types():
                                ats.append(AnalysisType(a))
                        except Exception:
                            pass
                        job = BatchJob(job_id, getattr(row, 'name', job_id), getattr(row, 'description', ''), ats, row.get_model_filter() or [], row.get_app_filter() or [], {})
                        job.status = row.status
                        job.total_tasks = row.total_tasks or 0
                        job.completed_tasks = row.completed_tasks or 0
                        job.failed_tasks = row.failed_tasks or 0
                        job.started_at = row.started_at
                        job.completed_at = row.completed_at
                        self.jobs.setdefault(job_id, job)
            except Exception:
                pass
        if not job:
            return ('text/plain', f'job_{job_id}_not_found.txt', b'Job not found')

        if fmt == 'json':
            import json as _json
            payload = job.to_dict()
            data_bytes = _json.dumps(payload, default=str, indent=2).encode('utf-8')
            self._persist_export_file(f'{job_id}.json', data_bytes)
            return ('application/json', f'{job_id}.json', data_bytes)
        if fmt == 'csv':
            import io, csv
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(['id','name','status','analysis_types','models','apps','total_tasks','completed','failed','cached'])
            writer.writerow([
                job.id,
                job.name,
                job.status.value,
                ';'.join(a.value for a in job.analysis_types),
                ';'.join(job.models),
                ';'.join(map(str, job.app_range)),
                job.total_tasks,
                job.completed_tasks,
                job.failed_tasks,
                job.cached_tasks,
            ])
            data_bytes = buf.getvalue().encode('utf-8')
            self._persist_export_file(f'{job_id}.csv', data_bytes)
            return ('text/csv', f'{job_id}.csv', data_bytes)
        if fmt in ('txt','summary'):
            lines = [
                f'Job {job.id}',
                f'Status: {job.status.value}',
                f'Progress: {job.completed_tasks}/{job.total_tasks}',
                f'Failed: {job.failed_tasks}',
                f'Models: {", ".join(job.models)}',
                f'Apps: {", ".join(map(str, job.app_range))}',
            ]
            data_bytes = '\n'.join(lines).encode('utf-8')
            self._persist_export_file(f'{job_id}.txt', data_bytes)
            return ('text/plain', f'{job_id}.txt', data_bytes)
        if fmt in ('xlsx','excel'):
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                import io as _io
                wb = Workbook()
                ws = wb.active
                ws.title = 'JobSummary'
                ws.append(['Field','Value'])
                meta_rows = [
                    ('Job ID', job.id),
                    ('Name', job.name),
                    ('Status', job.status.value),
                    ('Models', ', '.join(job.models)),
                    ('Apps', ', '.join(map(str, job.app_range))),
                    ('Analysis Types', ', '.join(a.value for a in job.analysis_types)),
                    ('Total Tasks', job.total_tasks),
                    ('Completed Tasks', job.completed_tasks),
                    ('Failed Tasks', job.failed_tasks),
                    ('Cached Tasks', job.cached_tasks),
                    ('Progress %', f"{job.progress_pct():.2f}"),
                ]
                for r in meta_rows:
                    ws.append(list(r))
                # Auto width
                for col in (1,2):
                    max_len = 0
                    for cell in ws[get_column_letter(col)]:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                    ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)
                bio = _io.BytesIO()
                wb.save(bio)
                xbytes = bio.getvalue()
                self._persist_export_file(f'{job_id}.xlsx', xbytes)
                return ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f'{job_id}.xlsx', xbytes)
            except Exception:  # pragma: no cover
                return ('text/plain', f'{job_id}.xlsx', b'Excel export failed')
        if fmt == 'pdf':
            try:
                from reportlab.lib.pagesizes import LETTER
                from reportlab.pdfgen import canvas
                import io as _io
                buf = _io.BytesIO()
                c = canvas.Canvas(buf, pagesize=LETTER)
                text = c.beginText(40, 750)
                text.setFont("Helvetica", 12)
                for line in [
                    f"Batch Job Report: {job.id}",
                    f"Name: {job.name}",
                    f"Status: {job.status.value}",
                    f"Progress: {job.completed_tasks}/{job.total_tasks}",
                    f"Failed: {job.failed_tasks}",
                    f"Models: {', '.join(job.models)}",
                    f"Apps: {', '.join(map(str, job.app_range))}",
                    f"Analysis Types: {', '.join(a.value for a in job.analysis_types)}",
                    f"Cached Tasks: {job.cached_tasks}",
                ]:
                    text.textLine(line)
                c.drawText(text)
                c.showPage()
                c.save()
                pdf_bytes = buf.getvalue()
                self._persist_export_file(f'{job_id}.pdf', pdf_bytes)
                return ('application/pdf', f'{job_id}.pdf', pdf_bytes)
            except Exception:  # pragma: no cover
                return ('application/pdf', f'{job_id}.pdf', b'%PDF-1.4\n% Placeholder PDF export for job')
        if fmt == 'ndjson':
            import json as _json
            line = job.to_dict()
            data_bytes = (_json.dumps(line)+'\n').encode('utf-8')
            self._persist_export_file(f'{job_id}.ndjson', data_bytes)
            return ('application/x-ndjson', f'{job_id}.ndjson', data_bytes)
        return ('text/plain', f'{job_id}.txt', b'Unsupported format')

    def generate_jobs_export(self, fmt: str = 'csv', include_in_memory_only: bool = True) -> tuple[str, str, bytes]:
        """Export summary of all known jobs."""
        fmt = (fmt or 'csv').lower()
        rows = list(self.jobs.values())
        if include_in_memory_only is False:
            try:  # pragma: no cover
                with get_session() as session:
                    db_rows = session.query(BatchAnalysis).order_by(BatchAnalysis.created_at.desc()).limit(200).all()
                    seen = {j.id for j in rows}
                    for r in db_rows:
                        if r.batch_id in seen:
                            continue
                        try:
                            ats = [AnalysisType(a) for a in r.get_analysis_types()]
                            bj = BatchJob(
                                r.batch_id,
                                getattr(r, 'name', r.batch_id),
                                getattr(r, 'description', ''),
                                ats,
                                r.get_model_filter() or [],
                                r.get_app_filter() or [],
                                {},
                            )
                            bj.status = r.status
                            bj.total_tasks = r.total_tasks or 0
                            bj.completed_tasks = r.completed_tasks or 0
                            bj.failed_tasks = r.failed_tasks or 0
                            bj.started_at = r.started_at
                            bj.completed_at = r.completed_at
                            rows.append(bj)
                        except Exception:
                            continue
            except Exception:
                pass

        if fmt == 'json':
            import json as _json
            data = [j.to_dict() for j in rows]
            data_bytes = _json.dumps(data, default=str, indent=2).encode('utf-8')
            self._persist_export_file('batch_jobs.json', data_bytes)
            return ('application/json', 'batch_jobs.json', data_bytes)
        if fmt == 'csv':
            import io, csv
            buf = io.StringIO()
            writer = csv.writer(buf)
            writer.writerow(['id','name','status','analysis_types','models','apps','total_tasks','completed','failed','cached','progress_pct'])
            for j in rows:
                writer.writerow([
                    j.id,
                    j.name,
                    j.status.value,
                    ';'.join(a.value for a in j.analysis_types),
                    ';'.join(j.models),
                    ';'.join(map(str, j.app_range)),
                    j.total_tasks,
                    j.completed_tasks,
                    j.failed_tasks,
                    j.cached_tasks,
                    f'{j.progress_pct():.2f}',
                ])
            data_bytes = buf.getvalue().encode('utf-8')
            self._persist_export_file('batch_jobs.csv', data_bytes)
            return ('text/csv', 'batch_jobs.csv', data_bytes)
        if fmt == 'pdf':
            try:
                from reportlab.lib.pagesizes import LETTER
                from reportlab.pdfgen import canvas
                import io as _io
                buf = _io.BytesIO()
                c = canvas.Canvas(buf, pagesize=LETTER)
                text = c.beginText(40, 750)
                text.setFont("Helvetica", 11)
                text.textLine("Batch Jobs Summary")
                text.textLine("")
                header = "ID | Name | Status | Progress"
                text.textLine(header)
                text.textLine('-' * len(header))
                for j in rows[:50]:
                    text.textLine(f"{j.id[:12]} | {j.name[:12]} | {j.status.value} | {j.completed_tasks}/{j.total_tasks}")
                c.drawText(text)
                c.showPage()
                c.save()
                pdf_bytes = buf.getvalue()
                self._persist_export_file('batch_jobs.pdf', pdf_bytes)
                return ('application/pdf', 'batch_jobs.pdf', pdf_bytes)
            except Exception:  # pragma: no cover
                return ('application/pdf', 'batch_jobs.pdf', b'%PDF-1.4\n% Placeholder PDF export for batch jobs list')
        if fmt in ('xlsx','excel'):
            try:
                import io as _io
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = 'BatchJobs'
                ws.append(['ID','Name','Status','Analysis Types','Models','Apps','Total Tasks','Completed','Failed','Cached','Progress %'])
                for j in rows:
                    ws.append([
                        j.id,
                        j.name,
                        j.status.value,
                        ';'.join(a.value for a in j.analysis_types),
                        ';'.join(j.models),
                        ';'.join(map(str, j.app_range)),
                        j.total_tasks,
                        j.completed_tasks,
                        j.failed_tasks,
                        j.cached_tasks,
                        f"{j.progress_pct():.2f}",
                    ])
                bio = _io.BytesIO()
                wb.save(bio)
                x_bytes = bio.getvalue()
                self._persist_export_file('batch_jobs.xlsx', x_bytes)
                return ('application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'batch_jobs.xlsx', x_bytes)
            except Exception:  # pragma: no cover
                return ('text/plain', 'batch_jobs.xlsx', b'Excel export not yet implemented')
        return ('text/plain', 'batch_jobs.txt', b'Unsupported format')

    def _persist_export_file(self, filename: str, data: bytes):  # pragma: no cover - side effect utility
        """Persist exported artifact with timestamp to avoid overwrites.

        Adds _YYYYMMDD_HHMMSS_%f before file extension. Silent on failure.
        """
        try:
            from datetime import datetime, timezone
            reports_dir = Paths.REPORTS_DIR
            reports_dir.mkdir(parents=True, exist_ok=True)
            stem, dot, ext = filename.rpartition('.')
            if not stem:  # filename without dot
                stem = filename
                ext = ''
            ts = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S_%f')
            ts_name = f"{stem}_{ts}.{ext}" if ext else f"{stem}_{ts}"
            target = reports_dir / ts_name
            with open(target, 'wb') as f:
                f.write(data)
        except Exception:
            pass

# Global singleton expected by importing code
batch_service = BatchAnalysisService()

__all__ = [
    'BatchAnalysisService',
    'BatchJob',
    'batch_service',
]
