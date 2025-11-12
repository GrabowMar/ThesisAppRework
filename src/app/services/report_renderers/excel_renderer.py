"""
Excel Report Renderer

Generates Excel reports using openpyxl with multiple sheets and charts.
"""
import logging
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def render_excel(report, data: Dict[str, Any], output_path: Path) -> None:
    """
    Render report as Excel workbook with multiple sheets.
    
    Args:
        report: Report model instance
        data: Report data dictionary
        output_path: Path where Excel file should be saved
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Route to appropriate renderer
        if report.report_type == 'app_analysis':
            _render_app_analysis_excel(wb, data)
        elif report.report_type == 'model_comparison':
            _render_model_comparison_excel(wb, data)
        elif report.report_type == 'tool_effectiveness':
            _render_tool_effectiveness_excel(wb, data)
        elif report.report_type == 'executive_summary':
            _render_executive_summary_excel(wb, data)
        else:
            _render_default_excel(wb, data)
        
        # Save workbook
        wb.save(output_path)
        
        logger.info(f"Rendered Excel report to {output_path}")
        
    except Exception as e:
        logger.error(f"Error rendering Excel report: {e}", exc_info=True)
        raise


def _render_app_analysis_excel(wb: Workbook, data: Dict[str, Any]) -> None:
    """Render app analysis data to Excel."""
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    _add_summary_sheet(ws_summary, data)
    
    # Findings sheet
    if data.get('findings'):
        ws_findings = wb.create_sheet("Findings")
        _add_findings_sheet(ws_findings, data['findings'])
    
    # Tools sheet
    if data.get('tools'):
        ws_tools = wb.create_sheet("Tools")
        _add_tools_sheet(ws_tools, data['tools'])


def _render_model_comparison_excel(wb: Workbook, data: Dict[str, Any]) -> None:
    """Render model comparison data to Excel."""
    # Comparison sheet
    ws_comparison = wb.create_sheet("Model Comparison")
    _add_model_comparison_sheet(ws_comparison, data)
    
    # Aggregated stats
    ws_stats = wb.create_sheet("Aggregated Stats")
    _add_aggregated_stats_sheet(ws_stats, data.get('aggregated', {}))


def _render_tool_effectiveness_excel(wb: Workbook, data: Dict[str, Any]) -> None:
    """Render tool effectiveness data to Excel."""
    ws_tools = wb.create_sheet("Tool Statistics")
    _add_tool_statistics_sheet(ws_tools, data.get('tools', {}))


def _render_executive_summary_excel(wb: Workbook, data: Dict[str, Any]) -> None:
    """Render executive summary data to Excel."""
    ws_exec = wb.create_sheet("Executive Summary")
    _add_executive_summary_sheet(ws_exec, data)


def _render_default_excel(wb: Workbook, data: Dict[str, Any]) -> None:
    """Render generic data to Excel."""
    ws = wb.create_sheet("Data")
    ws.append(["Key", "Value"])
    
    for key, value in data.items():
        ws.append([str(key), str(value)])
    
    _apply_header_style(ws, 1)


def _add_summary_sheet(ws, data: Dict[str, Any]) -> None:
    """Add summary information sheet."""
    # Title
    ws.append(["Analysis Report Summary"])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(size=16, bold=True)
    ws.append([])
    
    # Basic info
    ws.append(["Model", data.get('model_slug', 'N/A')])
    ws.append(["App Number", data.get('app_number', 'N/A')])
    ws.append(["Task ID", data.get('task_id', 'N/A')])
    ws.append(["Generated", data.get('timestamp', 'N/A')])
    ws.append([])
    
    # Analysis summary
    analysis = data.get('analysis', {})
    summary = data.get('summary', {})
    
    ws.append(["Summary Statistics"])
    ws['A' + str(ws.max_row)].font = Font(bold=True)
    
    ws.append(["Total Findings", summary.get('total_findings', 0)])
    ws.append(["Services Executed", summary.get('services_executed', 0)])
    ws.append(["Tools Executed", summary.get('tools_executed', 0)])
    ws.append([])
    
    # Severity breakdown
    severity = summary.get('severity_breakdown', {})
    if severity:
        ws.append(["Severity Breakdown"])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        
        for sev, count in severity.items():
            ws.append([sev.capitalize(), count])
    
    _apply_formatting(ws)


def _add_findings_sheet(ws, findings: List[Dict[str, Any]]) -> None:
    """Add findings table sheet."""
    # Headers
    headers = ["ID", "Tool", "Severity", "Category", "File", "Line", "Title", "Description"]
    ws.append(headers)
    _apply_header_style(ws, 1)
    
    # Data rows
    for finding in findings:
        file_info = finding.get('file', {})
        message = finding.get('message', {})
        
        ws.append([
            finding.get('id', ''),
            finding.get('tool', ''),
            finding.get('severity', ''),
            finding.get('category', ''),
            file_info.get('path', ''),
            file_info.get('line_start', ''),
            message.get('title', ''),
            message.get('description', '')[:100]  # Truncate long descriptions
        ])
    
    _apply_formatting(ws)
    
    # Add filters
    ws.auto_filter.ref = ws.dimensions


def _add_tools_sheet(ws, tools: Dict[str, Any]) -> None:
    """Add tools execution summary sheet."""
    # Headers
    headers = ["Tool", "Status", "Total Issues", "Duration (s)", "Executed"]
    ws.append(headers)
    _apply_header_style(ws, 1)
    
    # Data rows
    for tool_name, tool_data in tools.items():
        ws.append([
            tool_name,
            tool_data.get('status', 'unknown'),
            tool_data.get('total_issues', 0),
            tool_data.get('duration_seconds', 0),
            'Yes' if tool_data.get('executed') else 'No'
        ])
    
    _apply_formatting(ws)


def _add_model_comparison_sheet(ws, data: Dict[str, Any]) -> None:
    """Add model comparison sheet."""
    models = data.get('models', [])
    
    # Headers
    headers = ["Model", "Task ID", "Total Findings", "Completed At"]
    ws.append(headers)
    _apply_header_style(ws, 1)
    
    # Data rows
    for model in models:
        ws.append([
            model.get('model_slug', ''),
            model.get('task_id', ''),
            model.get('findings_count', 0),
            model.get('completed_at', '')
        ])
    
    _apply_formatting(ws)
    
    # Add chart if we have data
    if len(models) > 0:
        _add_comparison_chart(ws, len(models))


def _add_aggregated_stats_sheet(ws, aggregated: Dict[str, Any]) -> None:
    """Add aggregated statistics sheet."""
    ws.append(["Metric", "Value"])
    _apply_header_style(ws, 1)
    
    for key, value in aggregated.items():
        ws.append([key.replace('_', ' ').title(), value])
    
    _apply_formatting(ws)


def _add_tool_statistics_sheet(ws, tools: Dict[str, Any]) -> None:
    """Add tool statistics sheet."""
    # Headers
    headers = ["Tool", "Total Runs", "Success Rate %", "Avg Duration", "Total Findings", "Avg Findings/Run"]
    ws.append(headers)
    _apply_header_style(ws, 1)
    
    # Data rows
    for tool_name, stats in tools.items():
        ws.append([
            tool_name,
            stats.get('total_runs', 0),
            round(stats.get('success_rate', 0), 2),
            round(stats.get('avg_duration', 0), 2),
            stats.get('total_findings', 0),
            round(stats.get('avg_findings_per_run', 0), 2)
        ])
    
    _apply_formatting(ws)


def _add_executive_summary_sheet(ws, data: Dict[str, Any]) -> None:
    """Add executive summary sheet."""
    summary = data.get('summary', {})
    
    ws.append(["Executive Summary"])
    ws.merge_cells('A1:B1')
    ws['A1'].font = Font(size=16, bold=True)
    ws.append([])
    
    ws.append(["Total Apps Generated", summary.get('total_apps_generated', 0)])
    ws.append(["Total Analyses Run", summary.get('total_analyses_run', 0)])
    ws.append(["Total Analyses Completed", summary.get('total_analyses_completed', 0)])
    ws.append(["Total Findings", summary.get('total_findings', 0)])
    ws.append(["Total Models", summary.get('total_models', 0)])
    ws.append([])
    
    # Severity breakdown
    severity = summary.get('severity_breakdown', {})
    if severity:
        ws.append(["Severity Breakdown"])
        ws['A' + str(ws.max_row)].font = Font(bold=True)
        
        for sev, count in severity.items():
            ws.append([sev.capitalize(), count])
    
    _apply_formatting(ws)


def _apply_header_style(ws, row: int) -> None:
    """Apply header styling to a row."""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _apply_formatting(ws) -> None:
    """Apply general formatting to worksheet."""
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width


def _add_comparison_chart(ws, num_models: int) -> None:
    """Add a comparison chart to the sheet."""
    try:
        chart = BarChart()
        chart.title = "Findings by Model"
        chart.x_axis.title = "Model"
        chart.y_axis.title = "Total Findings"
        
        data = Reference(ws, min_col=3, min_row=1, max_row=num_models + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=num_models + 1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, f"F2")
    except Exception as e:
        logger.warning(f"Could not add chart: {e}")
